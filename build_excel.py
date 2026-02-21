"""
ClearMetric Stock Options Calculator (ISO vs NSO) — Premium Excel Template
Product for Gumroad ($16.99)

3 Sheets:
  1. Options Calculator — inputs + ISO/NSO comparison + net proceeds
  2. Exercise Scenarios — compare exercising at different prices/times
  3. How To Use — ISO vs NSO explained, holding periods, AMT basics

Design: Steel Blue palette (#2C3E6B primary, #1B2A4A dark, #D4E6F1 input)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# ============================================================
# DESIGN SYSTEM — Steel Blue
# ============================================================
PRIMARY = "2C3E6B"
DARK = "1B2A4A"
WHITE = "FFFFFF"
INPUT_COLOR = "D4E6F1"
LIGHT_GRAY = "F5F6FA"
MED_GRAY = "D5D8DC"
DARK_GRAY = "5D6D7E"
LIGHT_BLUE = "E8EEF4"
ACCENT = "3498DB"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, color="B0BEC5", italic=True)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=11, color="2C3E50")
FONT_INPUT = Font(name="Calibri", size=12, color=DARK, bold=True)
FONT_VALUE = Font(name="Calibri", size=11, color="2C3E50")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=DARK)
FONT_SMALL = Font(name="Calibri", size=9, color=DARK_GRAY, italic=True)

FILL_PRIMARY = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_COLOR, end_color=INPUT_COLOR, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

THIN = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def header_bar(ws, row, c1, c2, text, fill=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = FONT_SECTION
    cell.fill = fill or FILL_PRIMARY
    cell.alignment = ALIGN_C
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill or FILL_PRIMARY
        ws.cell(row=row, column=c).border = THIN


def label_input(ws, row, lc, vc, label, value=None, fmt=None):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=value)
    cv.font = FONT_INPUT
    cv.fill = FILL_INPUT
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def label_calc(ws, row, lc, vc, label, formula, fmt=None, bold=False):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=formula)
    cv.font = FONT_BOLD if bold else FONT_VALUE
    cv.fill = FILL_WHITE
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def cols(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ============================================================
# SHEET 1: Options Calculator
# ============================================================
def build_calculator(ws):
    ws.title = "Options Calculator"
    ws.sheet_properties.tabColor = PRIMARY
    cols(ws, {"A": 2, "B": 28, "C": 14, "D": 4, "E": 28, "F": 14, "G": 4, "H": 28, "I": 14, "J": 2})

    for r in range(1, 60):
        for c in range(1, 11):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 9):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:I1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:I2")
    ws.row_dimensions[2].height = 38
    title = ws.cell(row=2, column=2, value="STOCK OPTIONS CALCULATOR (ISO vs NSO)")
    title.font = FONT_TITLE
    title.alignment = ALIGN_C
    ws.merge_cells("B3:I3")
    ws.row_dimensions[3].height = 22
    sub = ws.cell(
        row=3, column=2,
        value="Enter your numbers in the light blue cells. ISO and NSO paths calculate side by side.",
    )
    sub.font = FONT_SUBTITLE
    sub.alignment = ALIGN_C

    # ===== INPUTS =====
    header_bar(ws, 5, 2, 3, "INPUTS")
    label_input(ws, 6, 2, 3, "Number of options", 10000, "#,##0")
    label_input(ws, 7, 2, 3, "Strike price ($)", 5, "$#,##0.00")
    label_input(ws, 8, 2, 3, "Current FMV ($)", 25, "$#,##0.00")
    label_input(ws, 9, 2, 3, "Vested % (0-100)", 50, "0%")
    label_input(ws, 10, 2, 3, "W-2 salary ($)", 150000, "$#,##0")
    label_input(ws, 11, 2, 3, "Filing: 1=Single 2=MFJ", 1, "0")
    label_input(ws, 12, 2, 3, "State tax rate (e.g. 0.093)", 0.093, "0.0%")
    label_input(ws, 13, 2, 3, "Expected exit price ($)", 50, "$#,##0.00")
    label_input(ws, 14, 2, 3, "Holding period met? (1=Yes)", 1, "0")

    # Vested options
    ws.cell(row=15, column=2, value="Vested options (auto)").font = FONT_LABEL
    ws.cell(row=15, column=2).fill = FILL_GRAY
    ws.cell(row=15, column=2).border = THIN
    ws.cell(row=15, column=3, value="=C6*C9/100")
    ws.cell(row=15, column=3).font = FONT_VALUE
    ws.cell(row=15, column=3).fill = FILL_WHITE
    ws.cell(row=15, column=3).number_format = "#,##0"
    ws.cell(row=15, column=3).border = THIN

    # ===== ISO COLUMN (E-F) =====
    header_bar(ws, 5, 5, 6, "ISO", FILL_DARK)
    label_calc(ws, 6, 5, 6, "Spread per share", "=C8-C7", "$#,##0.00")
    label_calc(ws, 7, 5, 6, "Total spread", "=F6*C15", "$#,##0")
    label_calc(ws, 8, 5, 6, "Cash to exercise", "=C7*C15", "$#,##0", bold=True)
    label_calc(ws, 9, 5, 6, "AMT at exercise (est)", "=F7*0.26", "$#,##0")
    label_calc(ws, 10, 5, 6, "LTCG at sale (est 20%)", "=(C13-C7)*C15*0.20", "$#,##0")
    label_calc(ws, 11, 5, 6, "State tax at sale", "=(C13-C7)*C15*C12", "$#,##0")
    label_calc(ws, 12, 5, 6, "ISO Total Tax", "=F9+F10+F11", "$#,##0")
    label_calc(ws, 13, 5, 6, "Gross sale proceeds", "=C13*C15", "$#,##0")
    label_calc(ws, 14, 5, 6, "ISO Net Proceeds", "=F13-F8-F12", "$#,##0", bold=True)

    # ===== NSO COLUMN (H-I) =====
    header_bar(ws, 5, 8, 9, "NSO", FILL_DARK)
    label_calc(ws, 6, 8, 9, "Spread per share", "=C8-C7", "$#,##0.00")
    label_calc(ws, 7, 8, 9, "Total spread", "=I6*C15", "$#,##0")
    label_calc(ws, 8, 8, 9, "Cash to exercise", "=C7*C15", "$#,##0", bold=True)
    label_calc(ws, 9, 8, 9, "Ordinary income tax (est 32%)", "=I7*0.32", "$#,##0")
    label_calc(ws, 10, 8, 9, "State tax at exercise", "=I7*C12", "$#,##0")
    label_calc(ws, 11, 8, 9, "LTCG at sale (est 20%)", "=(C13-C8)*C15*0.20", "$#,##0")
    label_calc(ws, 12, 8, 9, "State tax at sale", "=(C13-C8)*C15*C12", "$#,##0")
    label_calc(ws, 13, 8, 9, "NSO Total Tax", "=I9+I10+I11+I12", "$#,##0")
    label_calc(ws, 14, 8, 9, "Gross sale proceeds", "=C13*C15", "$#,##0")
    label_calc(ws, 15, 8, 9, "NSO Net Proceeds", "=I14-I8-I13", "$#,##0", bold=True)

    # Verdict row
    ws.merge_cells("B17:I17")
    verdict = ws.cell(
        row=17, column=2,
        value='=IF(F14>I15,"ISO wins by $"&TEXT(F14-I15,"#,##0"),"NSO wins by $"&TEXT(I15-F14,"#,##0"))',
    )
    verdict.font = Font(name="Calibri", size=14, bold=True, color=DARK)
    verdict.fill = FILL_LIGHT
    verdict.border = THIN
    verdict.alignment = ALIGN_C

    ws.protection.sheet = True
    input_cells = [(6, 3), (7, 3), (8, 3), (9, 3), (10, 3), (11, 3), (12, 3), (13, 3), (14, 3)]
    for r, c in input_cells:
        ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 2: Exercise Scenarios
# ============================================================
def build_scenarios(wb):
    ws = wb.create_sheet("Exercise Scenarios")
    ws.sheet_properties.tabColor = ACCENT
    cols(ws, {"A": 2, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 2})

    for r in range(1, 45):
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    ws.merge_cells("B1:F2")
    ws.cell(row=1, column=2, value="EXERCISE SCENARIOS").font = FONT_TITLE
    ws.cell(row=1, column=2).fill = FILL_DARK
    ws.cell(row=1, column=2).alignment = ALIGN_C
    for r in range(1, 3):
        for c in range(2, 7):
            ws.cell(row=r, column=c).fill = FILL_DARK

    ws.merge_cells("B3:F3")
    ws.cell(row=3, column=2, value="Compare exercising at different exit prices. Uses inputs from Options Calculator.").font = FONT_SMALL
    ws.cell(row=3, column=2).alignment = ALIGN_L

    # Headers
    headers = ["Exit Price", "Gross Sale", "ISO Tax", "ISO Net", "NSO Tax", "NSO Net"]
    for col, label in enumerate(headers, start=2):
        c = ws.cell(row=5, column=col, value=label)
        c.font = FONT_HEADER
        c.fill = FILL_PRIMARY
        c.border = THIN
        c.alignment = ALIGN_C

    sh = "'Options Calculator'!"
    for i in range(20):
        row = 6 + i
        price = 10 + i * 5  # $10, $15, $20, ... $105
        ws.cell(row=row, column=2, value=price)
        ws.cell(row=row, column=2).number_format = "$#,##0"
        ws.cell(row=row, column=2).font = FONT_VALUE
        ws.cell(row=row, column=2).border = THIN

        # Gross = price * vested
        ws.cell(row=row, column=3, value=f"=B{row}*{sh}C15")
        ws.cell(row=row, column=3).number_format = "$#,##0"
        ws.cell(row=row, column=3).font = FONT_VALUE
        ws.cell(row=row, column=3).border = THIN

        # ISO tax: AMT on spread at exercise + LTCG + state at sale
        ws.cell(row=row, column=4, value=f"=({sh}C8-{sh}C7)*{sh}C15*0.26+MAX(0,B{row}-{sh}C7)*{sh}C15*0.20+MAX(0,B{row}-{sh}C7)*{sh}C15*{sh}C12")
        ws.cell(row=row, column=4).number_format = "$#,##0"
        ws.cell(row=row, column=4).font = FONT_VALUE
        ws.cell(row=row, column=4).border = THIN

        # ISO net
        ws.cell(row=row, column=5, value=f"=C{row}-D{row}-{sh}F8")
        ws.cell(row=row, column=5).number_format = "$#,##0"
        ws.cell(row=row, column=5).font = FONT_VALUE
        ws.cell(row=row, column=5).border = THIN

        # NSO tax (simplified)
        ws.cell(row=row, column=6, value=f"=({sh}C8-{sh}C7)*{sh}C15*0.32+({sh}C8-{sh}C7)*{sh}C15*{sh}C12+MAX(0,B{row}-{sh}C8)*{sh}C15*0.20+MAX(0,B{row}-{sh}C8)*{sh}C15*{sh}C12")
        ws.cell(row=row, column=6).number_format = "$#,##0"
        ws.cell(row=row, column=6).font = FONT_VALUE
        ws.cell(row=row, column=6).border = THIN

        # NSO net
        ws.cell(row=row, column=7, value=f"=C{row}-F{row}-{sh}I8")
        ws.cell(row=row, column=7).number_format = "$#,##0"
        ws.cell(row=row, column=7).font = FONT_VALUE
        ws.cell(row=row, column=7).border = THIN

    ws.protection.sheet = True
    for r in range(6, 26):
        ws.cell(row=r, column=2).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 3: How To Use
# ============================================================
def build_instructions(wb):
    ws = wb.create_sheet("How To Use")
    ws.sheet_properties.tabColor = DARK_GRAY
    cols(ws, {"A": 3, "B": 90})

    ws.merge_cells("A1:B2")
    c = ws.cell(row=1, column=1, value="HOW TO USE THE STOCK OPTIONS CALCULATOR")
    c.font = FONT_TITLE
    c.fill = FILL_DARK
    c.alignment = ALIGN_C
    for r in range(1, 3):
        for co in range(1, 3):
            ws.cell(row=r, column=co).fill = FILL_DARK

    sections = [
        ("QUICK START", [
            "1. Open the 'Options Calculator' tab and enter your numbers in the LIGHT BLUE cells",
            "2. Number of options, strike price, current FMV, vested %, W-2 salary",
            "3. Filing status (1=Single, 2=Married Filing Jointly), state tax rate",
            "4. Expected exit price and whether holding period will be met",
            "5. ISO and NSO columns update automatically with tax estimates",
            "6. Use 'Exercise Scenarios' to compare different exit prices",
        ]),
        ("ISO (INCENTIVE STOCK OPTIONS)", [
            "No regular income tax at exercise — but AMT may apply on the spread (FMV - strike)",
            "AMT preference: spread is added to AMT income. Exemption phases out at high income",
            "At sale: LTCG if you hold 1 year from exercise AND 2 years from grant (qualifying disposition)",
            "Disqualifying disposition: ordinary income on spread at exercise + capital gains on appreciation",
            "AMT credit: If you pay AMT, you may get credit in future years when regular tax > AMT",
        ]),
        ("NSO (NON-QUALIFIED STOCK OPTIONS)", [
            "Ordinary income tax at exercise on the spread (FMV - strike) × shares",
            "Employer typically withholds (FICA + income tax) at exercise",
            "At sale: capital gains on appreciation from exercise FMV to sale price",
            "LTCG if held > 1 year from exercise; STCG (ordinary rates) if sold sooner",
        ]),
        ("HOLDING PERIODS", [
            "ISO qualifying: 1 year from EXERCISE + 2 years from GRANT",
            "NSO LTCG: 1 year from EXERCISE on the appreciation portion",
            "Early exercise with 83(b): Can start holding period at grant (restricted stock only)",
            "Plan ahead — selling too soon can cost you 10–20% in extra tax",
        ]),
        ("AMT BASICS", [
            "AMT is a parallel tax system. You pay the higher of regular tax or AMT",
            "ISO spread at exercise is an AMT preference item — no regular tax, but AMT may apply",
            "2026 exemption: ~$90K single, ~$140K MFJ. Phase-out above ~$500K / $1M",
            "AMT rates: 26% on first portion, 28% above. Consult Form 6251 for your situation",
        ]),
        ("IMPORTANT NOTES", [
            "This is an estimator only — consult a CPA for your specific situation",
            "Tax rates used are simplified (effective rates). Actual brackets vary",
            "State tax treatment may differ. Some states conform to federal AMT",
            "© 2026 ClearMetric. For educational use only. Not financial or tax advice.",
        ]),
    ]

    r = 4
    for title, items in sections:
        ws.cell(row=r, column=2, value=title).font = Font(name="Calibri", size=12, bold=True, color=DARK)
        ws.cell(row=r, column=2).fill = FILL_LIGHT
        ws.cell(row=r, column=2).border = THIN
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item).font = Font(name="Calibri", size=11, color="2C3E50")
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    print("Building Options Calculator sheet...")
    build_calculator(ws)

    print("Building Exercise Scenarios sheet...")
    build_scenarios(wb)

    print("Building How To Use sheet...")
    build_instructions(wb)

    wb.active = 0

    out = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "output",
        "ClearMetric-Stock-Options-Calculator.xlsx",
    )
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Size: {os.path.getsize(out) / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
